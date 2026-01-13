# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
from dataclasses import dataclass
from datetime import time
from io import BytesIO
from pathlib import Path

import openpyxl

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import mm

from PIL import Image, ImageDraw, ImageFont

APP_TITLE = "דף נחיתה – חישוב עלויות ומחירים"

# ---------------------------
# Helpers
# ---------------------------
def _clean_xlsx_path(p: str) -> str:
    # Handle accidental direction marks / invisible unicode
    return "".join(ch for ch in p if ch.isprintable())

def load_template_xlsx(uploaded_file) -> openpyxl.Workbook:
    wb = openpyxl.load_workbook(uploaded_file, data_only=False)
    return wb

def time_to_hours(t):
    if t is None:
        return 0.0
    if isinstance(t, time):
        return t.hour + t.minute/60 + t.second/3600
    # Streamlit time_input returns datetime.time
    return float(getattr(t, "hour", 0)) + float(getattr(t, "minute", 0))/60 + float(getattr(t, "second", 0))/3600

def mround(value: float, multiple: float) -> float:
    if multiple == 0:
        return value
    return round(value / multiple) * multiple

def currency(n: float) -> str:
    return f"{n:,.0f} ₪".replace(",", ",")

def currency2(n: float) -> str:
    return f"{n:,.2f} ₪".replace(",", ",")

def discount_factor(qty: int) -> float:
    if 1 < qty <= 30:
        return 0.9
    if 30 < qty <= 100:
        return 0.8
    if qty > 100:
        return 0.75
    return 1.0

def read_rates_from_sheet(wb: openpyxl.Workbook):
    ws = wb.active
    # Materials table in rows 6-9: C name, D price per kg
    materials = {}
    for r in range(6, 10):
        name = ws.cell(r, 3).value
        per_kg = ws.cell(r, 4).value
        if name:
            materials[str(name)] = float(per_kg or 0.0)
    # Work rates in rows 13-15: C name, D price per hour
    work = {}
    for r in range(13, 16):
        name = ws.cell(r, 3).value
        per_h = ws.cell(r, 4).value
        if name:
            work[str(name)] = float(per_h or 0.0)

    # Add-ons prices in rows 18-20: C name, D price
    addons = {}
    for r in range(18, 21):
        name = ws.cell(r, 3).value
        price = ws.cell(r, 4).value
        if name:
            addons[str(name)] = float(price or 0.0)
    return materials, work, addons

@dataclass
class Inputs:
    project_name: str
    material_lines: list  # list of dicts: {"חומר": str, "גרמים": float}
    modeling_time: time
    printing_time: time
    assembly_time: time
    magnets_qty: int
    led_single_qty: int
    led_desk_qty: int
    units_qty: int

def compute(inputs: Inputs, materials_per_kg: dict, work_per_h: dict, addons_price: dict):
    per_g = {k: v/1000 for k, v in materials_per_kg.items()}

    # Material costs
    materials_rows = []
    mat_total = 0.0
    for i, line in enumerate(inputs.material_lines, start=1):
        mat = line["חומר"]
        grams = float(line.get("גרמים", 0) or 0)
        cost = per_g.get(mat, 0.0) * grams
        materials_rows.append({"סעיף": f"חומר {i}", "פירוט": f"{mat} — {grams:.0f} גרם", "עלות": cost})
        mat_total += cost

    # Labor costs
    modeling_h = time_to_hours(inputs.modeling_time)
    printing_h = time_to_hours(inputs.printing_time)
    assembly_h = time_to_hours(inputs.assembly_time)

    modeling_cost = work_per_h.get("מידול", 0.0) * modeling_h
    printing_cost = work_per_h.get("הדפסה", 0.0) * printing_h
    assembly_cost = work_per_h.get("הרכבה", 0.0) * assembly_h

    labor_rows = [
        {"סעיף": "עבודה", "פירוט": f"מידול — {modeling_h:.2f} שעות", "עלות": modeling_cost},
        {"סעיף": "עבודה", "פירוט": f"הדפסה — {printing_h:.2f} שעות", "עלות": printing_cost},
        {"סעיף": "עבודה", "פירוט": f"הרכבה — {assembly_h:.2f} שעות", "עלות": assembly_cost},
    ]
    labor_total = modeling_cost + printing_cost + assembly_cost

    # Add-ons
    magnets_cost = addons_price.get("מגנטים (שקל/מגנט)", 0.0) * int(inputs.magnets_qty or 0)
    led_single_cost = addons_price.get("לד בודד", 0.0) * int(inputs.led_single_qty or 0)
    led_desk_cost = addons_price.get("לד שולחני", 0.0) * int(inputs.led_desk_qty or 0)

    addon_rows = [
        {"סעיף": "תוספות", "פירוט": f"מגנטים — {int(inputs.magnets_qty or 0)} יח׳", "עלות": magnets_cost},
        {"סעיף": "תוספות", "פירוט": f"לד בודד — {int(inputs.led_single_qty or 0)} יח׳", "עלות": led_single_cost},
        {"סעיף": "תוספות", "פירוט": f"לד שולחני — {int(inputs.led_desk_qty or 0)} יח׳", "עלות": led_desk_cost},
    ]
    addons_total = magnets_cost + led_single_cost + led_desk_cost

    # Pricing outputs (match sheet logic)
    cost_sum_all = mat_total + labor_total + addons_total
    unit_price_excl_modeling = (cost_sum_all - modeling_cost)
    qty = int(inputs.units_qty or 0)
    disc = discount_factor(qty)
    total = mround(modeling_cost + unit_price_excl_modeling * qty * disc, 5)

    rows = materials_rows + labor_rows + addon_rows
    df = pd.DataFrame(rows)
    return {
        "breakdown_df": df,
        "materials_total": mat_total,
        "labor_total": labor_total,
        "addons_total": addons_total,
        "modeling_cost": modeling_cost,
        "unit_price": unit_price_excl_modeling,
        "qty": qty,
        "discount": disc,
        "total": total,
        "project": inputs.project_name,
    }

def render_pdf(result: dict) -> bytes:
    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    width, height = A4

    # Minimal layout
    x = 18*mm
    y = height - 20*mm
    c.setFont("Helvetica-Bold", 18)
    c.drawRightString(width - x, y, "סיכום תמחור")
    y -= 10*mm
    c.setFont("Helvetica", 12)
    c.drawRightString(width - x, y, f"פרויקט: {result['project']}")
    y -= 8*mm
    c.drawRightString(width - x, y, f"תאריך: {pd.Timestamp.now().strftime('%d/%m/%Y')}")
    y -= 14*mm

    # Key numbers
    c.setFont("Helvetica-Bold", 12)
    c.drawRightString(width - x, y, f"מחיר יחידה (ללא מידול): {currency(result['unit_price'])}")
    y -= 7*mm
    c.drawRightString(width - x, y, f"כמות: {result['qty']}")
    y -= 7*mm
    c.drawRightString(width - x, y, f"הנחת כמות: {int((1-result['discount'])*100)}%")
    y -= 7*mm
    c.drawRightString(width - x, y, f"סה\"כ: {currency(result['total'])}")
    y -= 12*mm

    # Breakdown table (simple)
    c.setFont("Helvetica-Bold", 11)
    c.drawRightString(width - x, y, "פירוט עלויות")
    y -= 8*mm
    c.setFont("Helvetica", 10)
    # Headers
    c.drawString(x, y, "עלות")
    c.drawString(x + 35*mm, y, "פירוט")
    y -= 6*mm

    for _, r in result["breakdown_df"].iterrows():
        if y < 20*mm:
            c.showPage()
            y = height - 20*mm
            c.setFont("Helvetica", 10)
        cost = r["עלות"]
        detail = str(r["פירוט"])
        c.drawString(x, y, currency2(float(cost)))
        c.drawString(x + 35*mm, y, detail[:80])
        y -= 5.5*mm

    c.showPage()
    c.save()
    return buf.getvalue()

def render_image(result: dict) -> bytes:
    # Simple white image with black text
    W, H = 1200, 800
    img = Image.new("RGB", (W, H), "white")
    draw = ImageDraw.Draw(img)

    # Font: use default; if DejaVu exists, use it
    font_paths = [
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSansCondensed.ttf",
    ]
    font = None
    for fp in font_paths:
        if Path(fp).exists():
            font = ImageFont.truetype(fp, 38)
            font_small = ImageFont.truetype(fp, 26)
            font_tiny = ImageFont.truetype(fp, 22)
            break
    if font is None:
        font = ImageFont.load_default()
        font_small = font
        font_tiny = font

    x_right = W - 60
    y = 50
    draw.text((x_right, y), "סיכום תמחור", fill="black", font=font, anchor="ra")
    y += 70
    draw.text((x_right, y), f"פרויקט: {result['project']}", fill="black", font=font_small, anchor="ra")
    y += 40
    draw.text((x_right, y), f"מחיר יחידה (ללא מידול): {currency(result['unit_price'])}", fill="black", font=font_small, anchor="ra")
    y += 40
    draw.text((x_right, y), f"כמות: {result['qty']} | הנחה: {int((1-result['discount'])*100)}% | סה\"כ: {currency(result['total'])}", fill="black", font=font_small, anchor="ra")
    y += 60

    draw.text((x_right, y), "פירוט עלויות", fill="black", font=font_small, anchor="ra")
    y += 40

    # Table-like list
    for _, r in result["breakdown_df"].iterrows():
        detail = str(r["פירוט"])
        cost = currency2(float(r["עלות"]))
        draw.text((60, y), cost, fill="black", font=font_tiny, anchor="la")
        draw.text((x_right, y), detail, fill="black", font=font_tiny, anchor="ra")
        y += 30
        if y > H - 40:
            break

    out = BytesIO()
    img.save(out, format="PNG")
    return out.getvalue()

def write_back_to_xlsx(template_wb: openpyxl.Workbook, inputs: Inputs, result: dict) -> bytes:
    ws = template_wb.active

    # Project name (H5 in source)
    ws["H5"].value = inputs.project_name

    # Material selections and grams (H7..H10, I7..I10)
    base_row = 7
    for i, line in enumerate(inputs.material_lines):
        ws.cell(base_row + i, 8).value = line["חומר"]  # H
        ws.cell(base_row + i, 9).value = float(line.get("גרמים", 0) or 0)  # I

    # Times (I11..I13)
    ws["I11"].value = inputs.modeling_time
    ws["I12"].value = inputs.printing_time
    ws["I13"].value = inputs.assembly_time

    # Add-ons qty (I14..I16)
    ws["I14"].value = int(inputs.magnets_qty or 0)
    ws["I15"].value = int(inputs.led_single_qty or 0)
    ws["I16"].value = int(inputs.led_desk_qty or 0)

    # Units qty (J21)
    ws["J21"].value = int(inputs.units_qty or 0)

    # Note: formulas in the sheet won't auto-calc outside Excel.
    # We also write computed outputs into their cells for convenience:
    ws["J18"].value = float(result["modeling_cost"])
    ws["J20"].value = float(result["unit_price"])
    ws["J22"].value = float(result["discount"])
    ws["J24"].value = float(result["total"])

    out = BytesIO()
    template_wb.save(out)
    return out.getvalue()

# ---------------------------
# UI
# ---------------------------
st.set_page_config(page_title=APP_TITLE, layout="wide")

st.markdown(
    """
    <style>
    .block-container {padding-top: 1.2rem; padding-bottom: 2rem;}
    .stDownloadButton button {border-radius: 10px;}
    </style>
    """,
    unsafe_allow_html=True,
)

st.title(APP_TITLE)

uploaded = st.file_uploader("העלה קובץ אקסל (xlsx)", type=["xlsx"])

if not uploaded:
    st.info("העלה את הקובץ כדי להתחיל.")
    st.stop()

wb = load_template_xlsx(uploaded)
materials_per_kg, work_per_h, addons_price = read_rates_from_sheet(wb)

# Defaults from template (best-effort)
ws = wb.active
default_project = ws["H5"].value or "פרויקט"
default_lines = []
for r in range(7, 10):
    default_lines.append({
        "חומר": (ws.cell(r, 8).value or list(materials_per_kg.keys())[0]),
        "גרמים": float(ws.cell(r, 9).value or 0),
    })
default_model_time = ws["I11"].value or time(1,0)
default_print_time = ws["I12"].value or time(2,40)
default_assy_time = ws["I13"].value or time(0,30)
default_magnets = int(ws["I14"].value or 0)
default_led_single = int(ws["I15"].value or 0)
default_led_desk = int(ws["I16"].value or 0)
default_units = int(ws["J21"].value or 1)

left, right = st.columns([1.05, 1])

with left:
    st.subheader("קלטים")
    project_name = st.text_input("שם פרויקט", value=str(default_project))

    st.markdown("**חומרים**")
    material_names = list(materials_per_kg.keys())
    mat_lines = []
    for i in range(3):
        c1, c2 = st.columns([1,1])
        with c1:
            mat = st.selectbox(f"חומר {i+1}", material_names, index=material_names.index(default_lines[i]["חומר"]) if default_lines[i]["חומר"] in material_names else 0, key=f"mat{i}")
        with c2:
            grams = st.number_input(f"גרמים {i+1}", min_value=0.0, step=1.0, value=float(default_lines[i]["גרמים"]), key=f"grams{i}")
        mat_lines.append({"חומר": mat, "גרמים": grams})

    st.markdown("**עבודה**")
    c1, c2, c3 = st.columns(3)
    with c1:
        modeling_time = st.time_input("זמן מידול", value=default_model_time)
    with c2:
        printing_time = st.time_input("זמן הדפסה", value=default_print_time)
    with c3:
        assembly_time = st.time_input("זמן הרכבה", value=default_assy_time)

    st.markdown("**תוספות**")
    c1, c2, c3 = st.columns(3)
    with c1:
        magnets_qty = st.number_input("כמות מגנטים", min_value=0, step=1, value=int(default_magnets))
    with c2:
        led_single_qty = st.number_input("כמות לד בודד", min_value=0, step=1, value=int(default_led_single))
    with c3:
        led_desk_qty = st.number_input("כמות לד שולחני", min_value=0, step=1, value=int(default_led_desk))

    units_qty = st.number_input("כמות יחידות", min_value=0, step=1, value=int(default_units))

inputs = Inputs(
    project_name=project_name,
    material_lines=mat_lines,
    modeling_time=modeling_time,
    printing_time=printing_time,
    assembly_time=assembly_time,
    magnets_qty=magnets_qty,
    led_single_qty=led_single_qty,
    led_desk_qty=led_desk_qty,
    units_qty=units_qty,
)

result = compute(inputs, materials_per_kg, work_per_h, addons_price)

with right:
    st.subheader("תוצאות")
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("עלות מידול", currency(result["modeling_cost"]))
    k2.metric("מחיר יחידה (ללא מידול)", currency(result["unit_price"]))
    k3.metric("הנחת כמות", f"{int((1-result['discount'])*100)}%")
    k4.metric("סה\"כ", currency(result["total"]))

    st.dataframe(
        result["breakdown_df"].assign(עלות=result["breakdown_df"]["עלות"].map(currency2)),
        use_container_width=True,
        hide_index=True
    )

    st.markdown("---")
    st.markdown("**ייצוא**")

    pdf_bytes = render_pdf(result)
    img_bytes = render_image(result)
    xlsx_bytes = write_back_to_xlsx(wb, inputs, result)

    c1, c2, c3 = st.columns(3)
    with c1:
        st.download_button("הורד PDF", data=pdf_bytes, file_name=f"{project_name}_סיכום.pdf", mime="application/pdf")
    with c2:
        st.download_button("הורד תמונה (PNG)", data=img_bytes, file_name=f"{project_name}_סיכום.png", mime="image/png")
    with c3:
        st.download_button("הורד אקסל מעודכן", data=xlsx_bytes, file_name=f"{project_name}_מעודכן.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

st.caption("הערה: החישוב מתבצע בתוך האפליקציה (לא באמצעות חישוב נוסחאות אקסל), כדי שהייצוא ל-PDF/תמונה יהיה עקבי.")
